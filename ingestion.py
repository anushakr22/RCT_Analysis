"""
ingestion.py — Robust data ingestion for messy RCT Excel/CSV files.

Handles:
  • Multiple sheets — user picks one or merges all
  • Merged cells spanning pre/post columns
  • Multi-row headers (group label row + column name row)
  • Blank rows at the top before data starts
  • Totals / summary rows mixed into data
  • Wide format (pre/post as separate columns) → auto-melt to long
  • Long format — passes through untouched
"""

from __future__ import annotations
import re
import io
from dataclasses import dataclass, field
from typing import Any

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# Public result container
# ─────────────────────────────────────────────────────────────────────────────
@dataclass
class IngestionResult:
    df: pd.DataFrame                        # final cleaned dataframe
    sheet_used: str | list[str]             # sheet name(s) consumed
    format_detected: str                    # "long" | "wide"
    header_rows_skipped: int                # how many blank/label rows were dropped
    merged_cells_resolved: int              # count of merged-cell fills
    summary_rows_removed: int               # count of total/summary rows stripped
    wide_cols_melted: list[str]             # original wide columns that were melted
    warnings: list[str] = field(default_factory=list)
    col_rename_map: dict[str, str] = field(default_factory=dict)   # original→clean name


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
# Words that flag a totals/summary row
_SUMMARY_TOKENS = re.compile(
    r'\b(total|mean|average|avg|median|sum|n=|n =|subtotal|grand\s*total|'
    r'overall|aggregate|pooled|combined|all\s*participants?)\b',
    re.I,
)

# Words that flag a section-header / label row (not data)
_LABEL_TOKENS = re.compile(
    r'\b(pre[-\s]?intervention|post[-\s]?intervention|baseline|follow[- ]?up|'
    r'pre[-\s]?test|post[-\s]?test|time\s*point|session|wave|phase|'
    r'treatment\s*group|control\s*group|intervention|week\s*\d|'
    r'month\s*\d|visit\s*\d)\b',
    re.I,
)

# Patterns that look like pre/post column suffixes for wide detection
_WIDE_SUFFIX = re.compile(
    r'[_\-\s\.](pre|post|baseline|t0|t1|t2|t3|fu|follow[_\-]?up|'
    r'wk?\d+|week\d+|session\d+|wave\d+|v\d+|visit\d+)$',
    re.I,
)
_WIDE_PREFIX = re.compile(
    r'^(pre|post|baseline|t0|t1|t2|t3|fu|follow[_\-]?up|'
    r'wk?\d+|week\d+|session\d+|wave\d+|v\d+|visit\d+)[_\-\s\.]',
    re.I,
)


def _clean_col_name(name: Any) -> str:
    """Stringify and strip a column name."""
    s = str(name).strip()
    s = re.sub(r'\s+', ' ', s)
    return s


def _is_mostly_empty(row: pd.Series, thresh: float = 0.7) -> bool:
    n = len(row)
    if n == 0:
        return True
    empties = row.isna() | (row.astype(str).str.strip() == '')
    return empties.sum() / n >= thresh


def _row_is_summary(row: pd.Series) -> bool:
    """Return True if any cell in the row looks like a totals/summary label."""
    for val in row:
        if pd.isna(val):
            continue
        if _SUMMARY_TOKENS.search(str(val)):
            return True
    return False


def _row_is_section_label(row: pd.Series) -> bool:
    """Return True if a row is a section header (e.g. 'Pre-intervention')."""
    non_empty = [v for v in row if not pd.isna(v) and str(v).strip() != '']
    if not non_empty:
        return False
    # If ≤2 non-empty cells and they look like labels → section header
    if len(non_empty) <= 2:
        return any(_LABEL_TOKENS.search(str(v)) for v in non_empty)
    return False


# ─────────────────────────────────────────────────────────────────────────────
# Step 1 — read raw sheet(s) from Excel preserving merged cells
# ─────────────────────────────────────────────────────────────────────────────
def _read_raw_excel(file_bytes: bytes) -> dict[str, pd.DataFrame]:
    """
    Use openpyxl to read every sheet, unmerging cells by forward-filling the
    top-left value across the merged range.  Returns {sheet_name: DataFrame}
    with no header parsing (header=None) so we can inspect all rows.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheets: dict[str, pd.DataFrame] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # ── Resolve merged cells ──────────────────────────────────────────
        merged_count = 0
        for merge_range in list(ws.merged_cells.ranges):
            top_left_val = ws.cell(merge_range.min_row, merge_range.min_col).value
            ws.unmerge_cells(str(merge_range))
            for row in ws.iter_rows(
                min_row=merge_range.min_row, max_row=merge_range.max_row,
                min_col=merge_range.min_col, max_col=merge_range.max_col,
            ):
                for cell in row:
                    if cell.value is None:
                        cell.value = top_left_val
                        merged_count += 1

        # ── Read into DataFrame ───────────────────────────────────────────
        data = [[cell.value for cell in row] for row in ws.iter_rows()]
        if not data:
            continue
        df = pd.DataFrame(data)
        df.attrs['merged_cells_resolved'] = merged_count
        df.attrs['sheet_name'] = sheet_name
        sheets[sheet_name] = df

    return sheets


# ─────────────────────────────────────────────────────────────────────────────
# Step 2 — detect where the real header row is
# ─────────────────────────────────────────────────────────────────────────────
def _find_header_row(df: pd.DataFrame, max_scan: int = 15) -> tuple[int, int]:
    """
    Scan the first `max_scan` rows to find:
      - label_row: a row that looks like a section/group label (may be None)
      - header_row: the row index of the actual column names

    Returns (header_row_idx, n_rows_skipped).
    Heuristic: the header row is the first row where ≥40% of cells are
    non-empty strings that are NOT purely numeric, AND covers multiple columns.
    We skip rows that are:
      - mostly empty
      - a single merged label (≤2 distinct non-null values across many cols)
    """
    n = min(max_scan, len(df))
    for i in range(n):
        row = df.iloc[i]
        non_empty = [v for v in row if not pd.isna(v) and str(v).strip() != '']
        if not non_empty:
            continue

        # Skip rows that look like merged section labels:
        # few unique values but spread over many columns (due to merge fill)
        unique_vals = set(str(v).strip() for v in non_empty)
        if len(unique_vals) <= 2 and len(non_empty) >= 3:
            # Single label repeated across merged range → skip
            continue

        numeric_count = sum(1 for v in non_empty
                            if isinstance(v, (int, float)) or
                            re.match(r'^-?\d+(\.\d+)?$', str(v).strip()))
        str_count = len(non_empty) - numeric_count
        # Header row has mostly string labels across many columns
        if str_count >= max(2, 0.4 * len(row)):
            return i, i
    return 0, 0


# ─────────────────────────────────────────────────────────────────────────────
# Step 3 — collapse multi-row headers
# ─────────────────────────────────────────────────────────────────────────────
def _collapse_headers(df: pd.DataFrame, header_row: int) -> tuple[pd.DataFrame, list[str]]:
    """
    Detect multi-row header patterns and produce clean single-row column names.

    Two cases handled:
    A) Merged-section-label row (e.g. "Pre-Intervention" spanning cols C-E)
       followed by the real column name row → prefix ambiguous columns only.
    B) No label row above → just use the header row as-is.
    """
    # ── Find the nearest label row above header_row ─────────────────────────
    label_row_idx: int | None = None
    label_by_col: dict[int, str] = {}

    for check_row in range(header_row - 1, max(header_row - 4, -1), -1):
        if check_row < 0:
            break
        row = df.iloc[check_row]
        non_empty = [(i, v) for i, v in enumerate(row)
                     if not pd.isna(v) and str(v).strip() != '']
        if not non_empty:
            continue
        unique_vals = set(str(v).strip() for _, v in non_empty)
        # Treat as a merged label row: few unique values, ≥2 columns
        if len(unique_vals) <= 4 and len(non_empty) >= 2:
            # Forward-fill: each distinct label covers columns until next label
            last_label = ''
            for col_idx, val in enumerate(row):
                s = str(val).strip() if not pd.isna(val) else ''
                if s:
                    last_label = s
                if last_label:
                    label_by_col[col_idx] = last_label
            label_row_idx = check_row
        break

    # ── Build column names from header_row ───────────────────────────────────
    header = df.iloc[header_row]
    raw_names = [_clean_col_name(v) if not pd.isna(v) else f'col_{i}'
                 for i, v in enumerate(header)]

    # Only apply prefix when the same raw column name appears under multiple
    # different labels (truly ambiguous) or when raw name is blank/generic
    from collections import Counter
    name_count = Counter(raw_names)
    ambiguous = {n for n, c in name_count.items() if c > 1}

    cols = []
    seen: dict[str, int] = {}
    for col_idx, base in enumerate(raw_names):
        prefix = label_by_col.get(col_idx, '')
        # Prepend prefix only if:
        #   • The column name is ambiguous (same name under different labels), OR
        #   • The column name is a generic placeholder (col_N)
        generic = re.match(r'^col_\d+$', base)
        if prefix and (base in ambiguous or generic):
            candidate = f'{prefix}_{base}'
        else:
            candidate = base

        # Deduplicate
        if candidate in seen:
            seen[candidate] += 1
            candidate = f'{candidate}_{seen[candidate]}'
        else:
            seen[candidate] = 0
        cols.append(candidate)

    # Slice data rows (everything after the header row)
    df_out = df.iloc[header_row + 1:].copy()
    df_out.columns = cols
    df_out = df_out.reset_index(drop=True)
    return df_out, cols


# ─────────────────────────────────────────────────────────────────────────────
# Step 4 — remove blank, summary and section-label rows from data body
# ─────────────────────────────────────────────────────────────────────────────
def _remove_junk_rows(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Strip blank rows, totals rows, and embedded section-label rows."""
    keep = []
    removed = 0
    for idx, row in df.iterrows():
        if _is_mostly_empty(row):
            removed += 1
            continue
        if _row_is_summary(row):
            removed += 1
            continue
        if _row_is_section_label(row):
            removed += 1
            continue
        keep.append(idx)
    return df.loc[keep].reset_index(drop=True), removed


# ─────────────────────────────────────────────────────────────────────────────
# Step 5 — infer wide vs long, melt if wide
# ─────────────────────────────────────────────────────────────────────────────
def _detect_format(df: pd.DataFrame) -> tuple[str, list[str]]:
    """
    Heuristic: if several numeric columns share a common stem with
    pre/post/session suffixes (or prefixes), the file is wide format.

    Returns ("wide", [time_suffixed_columns]) or ("long", []).
    """
    cols = [str(c) for c in df.columns]
    # Check suffix pattern
    suffixed = [c for c in cols if _WIDE_SUFFIX.search(c)]
    prefixed = [c for c in cols if _WIDE_PREFIX.match(c)]
    candidates = suffixed if len(suffixed) >= len(prefixed) else prefixed

    if len(candidates) >= 2:
        # Confirm at least some of these are numeric
        numeric_cands = [c for c in candidates
                         if pd.to_numeric(df[c], errors='coerce').notna().mean() > 0.5]
        if len(numeric_cands) >= 2:
            return "wide", candidates
    return "long", []


def _extract_time_label(col: str) -> tuple[str, str]:
    """
    Given a column like 'pain_score_pre' or 'post_depression',
    return (stem, time_label).
    """
    m = _WIDE_SUFFIX.search(col)
    if m:
        stem = col[: m.start()]
        return stem.strip('_- '), m.group(1)
    m = _WIDE_PREFIX.match(col)
    if m:
        stem = col[m.end():]
        return stem.strip('_- '), m.group(1)
    return col, col


def _melt_wide_to_long(
    df: pd.DataFrame,
    wide_cols: list[str],
) -> tuple[pd.DataFrame, list[str]]:
    """
    Melt wide-format time columns to long format.

    Groups columns by stem (e.g. pain_pre + pain_post → pain, with
    a new 'time' column).  ID columns (non-time columns) are preserved.
    """
    id_cols = [c for c in df.columns if c not in wide_cols]

    # Group wide cols by stem
    stem_map: dict[str, list[tuple[str, str]]] = {}   # stem → [(col, time_label)]
    for c in wide_cols:
        stem, time_label = _extract_time_label(c)
        stem_map.setdefault(stem, []).append((c, time_label))

    # Build a long frame for each stem then merge
    long_frames = []
    for stem, pairs in stem_map.items():
        sub_cols = [p[0] for p in pairs]
        time_labels = [p[1] for p in pairs]
        melted = df[id_cols + sub_cols].melt(
            id_vars=id_cols,
            value_vars=sub_cols,
            var_name='_time_raw',
            value_name=stem,
        )
        # Map original column name → clean time label
        label_map = dict(zip(sub_cols, time_labels))
        melted['time'] = melted['_time_raw'].map(label_map)
        melted = melted.drop(columns=['_time_raw'])
        long_frames.append(melted)

    if not long_frames:
        return df, []

    # Merge all stems on id_cols + time
    result = long_frames[0]
    for frame in long_frames[1:]:
        merge_keys = id_cols + ['time']
        result = result.merge(frame, on=merge_keys, how='outer')

    return result, [p[0] for pairs in stem_map.values() for p in pairs]


# ─────────────────────────────────────────────────────────────────────────────
# Step 6 — type cleanup
# ─────────────────────────────────────────────────────────────────────────────
def _coerce_types(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in df.columns:
        # Try numeric coercion
        numeric = pd.to_numeric(df[col], errors='coerce')
        if numeric.notna().sum() / max(len(df), 1) > 0.6:
            df[col] = numeric
        else:
            df[col] = df[col].astype(str).str.strip()
            df[col] = df[col].replace({'nan': np.nan, 'None': np.nan, '': np.nan})
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Public API
# ─────────────────────────────────────────────────────────────────────────────
def get_sheet_names(file_bytes: bytes) -> list[str]:
    """Return list of sheet names from an xlsx file."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def get_sheet_preview(file_bytes: bytes, sheet_name: str, nrows: int = 8) -> pd.DataFrame:
    """
    Return a raw (no cleaning) preview of a sheet so the user can
    see what it looks like before committing.
    """
    raw_sheets = _read_raw_excel(file_bytes)
    if sheet_name not in raw_sheets:
        return pd.DataFrame()
    df = raw_sheets[sheet_name]
    return df.iloc[:nrows].fillna('').astype(str)


def ingest(
    file_bytes: bytes,
    filename: str,
    sheet_names: list[str] | None = None,   # None = all sheets; list = selected
    merge_sheets: bool = False,
) -> IngestionResult:
    """
    Full ingestion pipeline.

    Parameters
    ----------
    file_bytes   : raw bytes of the uploaded file
    filename     : original filename (used to detect csv vs xlsx)
    sheet_names  : which sheets to process (None = first/all)
    merge_sheets : if True, concatenate selected sheets after cleaning each

    Returns
    -------
    IngestionResult with .df holding the cleaned long-format DataFrame
    """
    warnings_out: list[str] = []
    total_merged = 0
    total_summary_removed = 0
    header_rows_skipped = 0
    format_detected = "long"
    wide_cols_melted: list[str] = []
    col_rename_map: dict[str, str] = {}

    # ── CSV path (simple) ─────────────────────────────────────────────────────
    if filename.lower().endswith('.csv'):
        df = pd.read_csv(io.BytesIO(file_bytes))
        df, removed = _remove_junk_rows(df)
        total_summary_removed = removed
        df = _coerce_types(df)
        fmt, wide_cols = _detect_format(df)
        if fmt == "wide":
            df, wide_cols_melted = _melt_wide_to_long(df, wide_cols)
            format_detected = "wide"
        return IngestionResult(
            df=df,
            sheet_used="(csv)",
            format_detected=format_detected,
            header_rows_skipped=0,
            merged_cells_resolved=0,
            summary_rows_removed=total_summary_removed,
            wide_cols_melted=wide_cols_melted,
            warnings=warnings_out,
            col_rename_map={},
        )

    # ── Excel path ────────────────────────────────────────────────────────────
    raw_sheets = _read_raw_excel(file_bytes)
    all_names = list(raw_sheets.keys())

    if not all_names:
        raise ValueError("No sheets found in this workbook.")

    # Determine which sheets to process
    if sheet_names is None:
        sheet_names = all_names[:1]   # default: first sheet
    target_sheets = [s for s in sheet_names if s in raw_sheets]
    if not target_sheets:
        raise ValueError(f"None of the requested sheets found. Available: {all_names}")

    cleaned_frames: list[pd.DataFrame] = []

    for sname in target_sheets:
        raw = raw_sheets[sname]
        total_merged += raw.attrs.get('merged_cells_resolved', 0)

        # Find header
        h_row, skipped = _find_header_row(raw)
        header_rows_skipped = max(header_rows_skipped, skipped)

        # Collapse multi-row headers
        df_clean, rename = _collapse_headers(raw, h_row)
        col_rename_map.update({r: r for r in rename})   # identity map for now

        # Remove junk rows
        df_clean, removed = _remove_junk_rows(df_clean)
        total_summary_removed += removed

        # Type coercion
        df_clean = _coerce_types(df_clean)

        if merge_sheets and len(target_sheets) > 1:
            df_clean['_source_sheet'] = sname

        cleaned_frames.append(df_clean)

    # Merge or single
    if merge_sheets and len(cleaned_frames) > 1:
        # Align columns: union of all columns
        all_cols = list(dict.fromkeys(
            c for f in cleaned_frames for c in f.columns
        ))
        aligned = [f.reindex(columns=all_cols) for f in cleaned_frames]
        df_final = pd.concat(aligned, ignore_index=True)
        if not all(f.columns.tolist() == cleaned_frames[0].columns.tolist()
                   for f in cleaned_frames):
            warnings_out.append(
                "Sheets had different columns — missing values filled with NaN "
                "where columns don't overlap."
            )
    else:
        df_final = cleaned_frames[0]

    # Detect wide vs long on final frame
    fmt, wide_cols = _detect_format(df_final)
    if fmt == "wide":
        df_final, wide_cols_melted = _melt_wide_to_long(df_final, wide_cols)
        format_detected = "wide"
        if wide_cols_melted:
            warnings_out.append(
                f"Wide format detected — {len(wide_cols_melted)} columns melted "
                f"into long format with a new 'time' column."
            )
    else:
        format_detected = "long"

    sheet_used = target_sheets if len(target_sheets) > 1 else target_sheets[0]

    return IngestionResult(
        df=df_final,
        sheet_used=sheet_used,
        format_detected=format_detected,
        header_rows_skipped=header_rows_skipped,
        merged_cells_resolved=total_merged,
        summary_rows_removed=total_summary_removed,
        wide_cols_melted=wide_cols_melted,
        warnings=warnings_out,
        col_rename_map=col_rename_map,
    )